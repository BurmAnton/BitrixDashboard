from django.shortcuts import render, redirect
from django.contrib import messages
from django.conf import settings
import django_filters
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import serializers
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework import viewsets
from rest_framework import status
from django_filters.rest_framework import DjangoFilterBackend
from .models import Contact, Organization, Projects, Region, FederalDistrict, OrganizationType, HistoryOrganization, ContactEmail, ContactPhone
from .forms import ContactImportFromExcel
from education_planner.models import ProfActivity, ROIV
from datetime import datetime
import openpyxl

def ExcelImportOrganization(form):
    if form.is_valid():
        excel_file = form.cleaned_data['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return False, f'Ошибка при открытии файла: {e}'

        created = 0
        updated = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            inn = row[0]
            name = row[1]
            full_name = row[2]
            type_name = row[3]
            federal_company = True if row[4] is not None else False
            region_name = row[5]
            parent_company_inn = row[6]

            OrgType = OrganizationType.objects.filter(name=type_name).first()
            region = Region.objects.filter(name=region_name).first()
            parent_company = Organization.objects.filter(inn=parent_company_inn) or None

            obj_history = Organization.objects.filter(inn=inn).first()
            if obj_history:       
                HistoryOrganization.objects.create(
                    organization=obj_history,
                    name=obj_history.name,
                    status='active',
                    date=datetime.now()
                )
                updated += 1
            else: created += 1

            Organization.objects.update_or_create(
                inn=inn,
                defaults={
                    'name': name,
                    'full_name': full_name,
                    'type': OrgType,
                    'federal_company': federal_company,
                    'region': region,
                    'parent_company': parent_company
                }
            )
        return True, f"Результат: Успешный импорт, Добавлено: {created}, Обновлено: {updated}"
    return False, f"Ошибка: Не корректная форма"

def ExcelImportContact(form):
    if form.is_valid():
        excel_file = form.cleaned_data['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return False, f'Ошибка при открытии файла: {e}'
        created_person = 0
        created_main = 0
        created_department = 0
        for SheetName in wb.sheetnames:
            if wb[SheetName] == "По сотрудникам":
                for row in ws.iter_rows(min_row=2, values_only=True):
                    inn = row[0]
                    last_name = row[1]
                    first_name = row[2]
                    middle_name = row[3]
                    position = row[4]
                    last_name_dat = row[5]
                    first_name_dat = row[6]
                    middle_name_dat = row[7]
                    position_dat = row[8]
                    manager = True if row[9] else False
                    phone = row[10]
                    email = row[11]
                    comment = row[12]


                    organization = Organization.objects.filter(inn=inn).first()
                    if organization:
                        created_person += 1
                        obj = Organization.objects.create(
                            organization=organization,
                            type='person',
                            defaults={
                                'first_name': first_name,
                                'last_name': last_name,
                                'middle_name': middle_name,
                                'position': position,
                                'first_name_dat': first_name_dat,
                                'last_name_dat': last_name_dat,
                                'middle_name_dat': middle_name_dat,
                                'position_dat': position_dat,
                                'manager': manager,
                                'comment': comment                        
                            }
                        )
                        if email:
                            ContactEmail.objects.create(
                                contact=obj,
                                email = email
                            )
                        if phone:
                            ContactPhone.objects.create(
                                contact=obj,
                                number=phone
                            )
            elif wb[SheetName] == "По отделам":
                for row in ws.iter_rows(min_row=2, values_only=True):
                    inn = row[0]
                    department_name = row[1]
                    phone = row[2]
                    email = row[3]
                    comment = row[4]

                    organization = Organization.objects.filter(inn=inn).first()
                    if organization:
                        created_department += 1
                        obj = Organization.objects.update_or_create(
                            organization=organization,
                            defaults={
                                'type': 'department',
                                'first_name': first_name,
                                'department_name': department_name,
                                'comment': comment                        
                            }
                        )
                        if email:
                            ContactEmail.objects.create(
                                contact=obj,
                                email = email
                            )
                        if phone:
                            ContactPhone.objects.create(
                                contact=obj,
                                number=phone
                            )
            elif wb[SheetName] == "Главный контакт":
                for row in ws.iter_rows(min_row=2, values_only=True):
                    inn = row[0]
                    phone = row[1]
                    email = row[2]
                    comment = row[3]


                    organization = Organization.objects.filter(inn=inn).first()
                    if organization:
                        created_main += 1
                        obj = Organization.objects.update_or_create(
                            organization=organization,
                            defaults={
                                'type': 'department',
                                'first_name': first_name,
                                'comment': comment                        
                            }
                        )
                        if email:
                            ContactEmail.objects.create(
                                contact=obj,
                                email = email
                            )
                        if phone:
                            ContactPhone.objects.create(
                                contact=obj,
                                number=phone
                            )
        return True, f'Импорт завершен успешно. Создано контактов: {created_person} сотрудников, {created_department} отделов, {created_main} основных контактов.'
    return False,f'Форма не валидна'
   
def ContactImport(request):
    if not request.user.is_authenticated:
        messages.warning(request, 'Для доступа к импорту необходимо войти в систему.')
        return redirect(f'{settings.LOGIN_URL}?next={request.path}')
    if request.method == 'POST':
        form = ContactImportFromExcel(request.POST,request.FILES)
        if form.is_valid():
            if form.cleaned_data['type'] == 'contacts':
                success, msg = ExcelImportContact(form)
                messages.success(request, msg) if success else messages.error(request, msg)
            elif form.cleaned_data['type'] == 'orgs':
                success, msg = ExcelImportOrganization(form)
                messages.success(request, msg) if success else messages.error(request, msg)
        else:
            print("Form errors:", form.errors)
            messages.error(request, "Не валидная форма")
    else:
        form = ContactImportFromExcel()

    return render(request,"contact_management/import_form.html", {'form': form})

def api_guide(request):
    from django.contrib.auth.models import User
    from rest_framework.authtoken.models import Token
    import requests
    import urllib.parse

    if not request.user.is_authenticated:
        messages.warning(request, 'Для доступа к REST API необходимо войти в систему.')
        return redirect(f'{settings.LOGIN_URL}?next={request.path}')
    if not request.user.is_staff and not request.user.is_superuser:
        return redirect('/')
    
    ChoiseFields = {}

    types = OrganizationType.objects.all()
    ChoiseFields.setdefault('type', [obj.name for obj in types])

    prof_activity = ProfActivity.objects.all()
    ChoiseFields.setdefault('prof_activity', [obj.name for obj in prof_activity])

    project = Projects.objects.all()
    ChoiseFields.setdefault('project', [obj.name for obj in project])

    region = Region.objects.all()
    ChoiseFields.setdefault('region', [obj.name for obj in region])

    fed_district = FederalDistrict.objects.all()
    ChoiseFields.setdefault('fed_district', [obj.name for obj in fed_district])

    organization = Organization.objects.all()
    ChoiseFields.setdefault('organization', [obj.name for obj in organization])

    login = User.objects.filter(username=request.user.username).first()
    tkn = Token.objects.filter(user=login).first()
    group = 'organization'
    if 'contact' in request.GET:
        group = 'contact'
    elif 'organization' in request.GET:
        group= 'organization'
    elif 'get_all' in request.GET:
        group= 'get_all'
        if 'model' in request.POST.dict():
            group = f'get_all/{request.POST.dict().get('model')}'
    url = f'http://{request.get_host()}/contacts/api/{group}' 

    headers = { 'Authorization': f'Token {tkn}' }

    params = {}
    for arg in request.POST.dict():
        value = request.POST.dict().get(arg)
        if arg != 'csrfmiddlewaretoken' and arg != 'apiLink' and value != '' and arg != 'model':
            params.setdefault(arg, value)
    
    if request.method == 'POST':
        response = requests.get(url,headers=headers, params=params)
    else:
        response = requests.get(url,headers=headers)
    import json
    try:
        jsn = json.dumps(response.json())
    except Exception as e:
        print(f"Не удалось преобразовать в json: {e}")
        jsn = str(response)
    context={
        'login': login,
        'token': tkn,
        'JsonReponse': jsn,
        'URL': str(urllib.parse.unquote(response.url)),
        'ChoiseFields': ChoiseFields
    }
    return render(request, f"contact_management/{'get_all' if 'get_all' in group else group}_api_form.html", context=context)

class ProjectsSerializer(serializers.ModelSerializer):
    """Сериалайзер для получения списка проектов"""
    class Meta:
        model = Projects
        fields = ["name"]

class OrganizationSerializer(serializers.ModelSerializer):
    """Сериалайзер для получения списка организаций"""
    type = serializers.CharField(source='type.name', read_only=True)
    region = serializers.CharField(source='region.name', read_only=True)
    prof_activity = serializers.CharField(source='prof_activity.name', read_only=True)
    fed_district = serializers.CharField(source='region.federalDistrict', read_only=True)
    projects = ProjectsSerializer(many=True, read_only=True)
    class Meta:
        model = Organization
        fields =[
            'inn',
            'name',
            'full_name',
            'type',
            'region',
            'federal_company',
            'fed_district',
            'prof_activity',
            'projects',
            'is_active',
            'created_at',
            'updated_at'
        ]
        
class OrganizationFilter(django_filters.FilterSet):
    """Фильтры API списков организаций"""
    type = django_filters.CharFilter(field_name='type__name', lookup_expr='exact')
    date = django_filters.CharFilter(field_name='created_at', lookup_expr='contains')
    
    prof_activity = django_filters.CharFilter(method='filter_prof_activity')
    prof_activity__contains = django_filters.CharFilter(method='filter_prof_activity_contains')

    project = django_filters.CharFilter(field_name='projects__name', lookup_expr='contains')

    region = django_filters.CharFilter(field_name='region__name', lookup_expr='exact')
    region__contains = django_filters.CharFilter(field_name='region__name', lookup_expr='contains')

    fed_district = django_filters.CharFilter(field_name='region__federalDistrict__name', lookup_expr='exact')
    fed_district__contains = django_filters.CharFilter(field_name='region__federalDistrict__name', lookup_expr='contains')

    federal = django_filters.BooleanFilter(field_name='federal_company')

    class Meta:
        model = Organization
        fields = []

    def filter_prof_activity(self, queryset, name, value):
        if not self.data.get("type") or self.data.get("type") != "РОИВ":
            return queryset
        return queryset.filter(prof_activity__name=value)

    def filter_prof_activity_contains(self, queryset, name, value):
        if not self.data.get("type") or self.data.get("type") != "РОИВ":
            return queryset
        return queryset.filter(prof_activity__name__contains=value)

class OrganizationViewSet(viewsets.ModelViewSet):
    """Viewset организаций, только чтение списка с учетом фильтров"""
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    filterset_class = OrganizationFilter
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    filter_backends = [DjangoFilterBackend]

    @action(detail=False, methods=["post"], url_path="add")
    def add_organization(self, request):
        serializer = self.get_serializer(data=request.data)
        print(serializer.is_valid())
        print(serializer.errors)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        projects = request.data.get("projects", [])
        prof_activity_names = request.data.get("prof_activity", [])
        # создаём контакт с найденной организацией
        obj = Organization.objects.create(
            inn=serializer.validated_data["inn"],
            name=serializer.validated_data.get("name", ""),
            full_name=serializer.validated_data.get("full_name", ""),
            type=OrganizationType.objects.filter(name=request.data.get("type", "")).first() or None,
            roiv=ROIV.objects.filter(name=request.data.get("roiv", "")).first() or None,
            region= Region.objects.filter(name=request.data.get("region", "")).first() or None,
            federal_company= serializer.validated_data.get("federal_company", False),
            is_active= serializer.validated_data.get("is_active", True),
            parent_company= Organization.objects.filter(inn=request.data.get("parent_company", "")).first() or None,
        )

        if projects is not None: 
            for name in projects:
                project = Projects.objects.filter(name=name).first()
                if project is not None:
                    obj.projects.add(project)
        if obj.type:
            if prof_activity_names is not None and obj.type.name == 'РОИВ':
                for name in prof_activity_names:
                    prof_activity = ProfActivity.objects.filter(name=name).first()
                    if prof_activity is not None:
                        obj.prof_activity.add(prof_activity)

        return Response(OrganizationSerializer(obj).data, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=["patch"], url_path="update")
    def update_organization(self, request):
        inn = request.data.get("inn")
        if not inn:
            return Response(
                {"inn": ["Обязательное поле."]},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obj = Organization.objects.filter(inn=inn).first()
        except Organization.DoesNotExist:
            return Response(
                {"detail": "Организация с таким ИНН не найдена."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        HistoryOrganization.objects.create(
            organization=obj,
            name=obj.name,
            status='active',
            date=datetime.now()
        )

        serializer = self.get_serializer(obj, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        prof_activity_names = request.data.get("prof_activity", None)
        projects = request.data.get("projects", None)
        # Обновляем поля модели
        for attr in request.data:
            if attr == "type":
                setattr(obj, attr, OrganizationType.objects.filter(name=request.data.get(attr, None)).first())
            elif attr == "roiv":
                setattr(obj, attr, ROIV.objects.filter(name=request.data.get(attr, None)).first())
            elif attr == "region":
                setattr(obj, attr, Region.objects.filter(name=request.data.get(attr, None)).first())
            elif attr == "parent_company":
                setattr(obj, attr, Organization.objects.filter(inn=request.data.get(attr, None)).first())
            else:
                try:
                    setattr(obj, attr, request.data.get(attr, None))
                except:
                    pass
        obj.save()
        # Обновляем ManyToMany prof_activity
        if obj.type:
            if prof_activity_names is not None and obj.type.name == 'РОИВ':
                for name in prof_activity_names:
                    prof_activity = ProfActivity.objects.filter(name=name).first()
                    if prof_activity is not None:
                        obj.prof_activity.add(prof_activity)

        # Обновляем ManyToMany projects
        if projects is not None: 
            for name in projects:
                project = Projects.objects.filter(name=name).first()
                if project is not None:
                    obj.projects.add(project)

        return Response(OrganizationSerializer(obj).data, status=status.HTTP_200_OK)

class ContactSerializer(serializers.ModelSerializer):
    """Сериалайзер для получения списка контактов """
    organization = serializers.CharField(source="organization.inn", read_only=True)

    class Meta:
        model = Contact
        fields = "__all__"
        # поля, которые можно указывать при создании/обновлении
        extra_kwargs = {
            "organization": {"required": True},
        }

    def to_representation(self, instance):
        data = super().to_representation(instance)

        if instance.type == "person":
            keep = [
                "id",
                "type",
                "comment",
                "current",
                "organization",
                "first_name",
                "last_name",
                "middle_name",
                "position",
                "first_name_dat",
                "last_name_dat",
                "middle_name_dat",
                "position_dat",
                "manager",
            ]
        elif instance.type == "department":
            keep = [
                "id",
                "type",
                "comment",
                "current",
                "organization",
                "department_name",
            ]
        else:
            keep = ["id", "type", "comment", "current", "organization"]

        return {k: v for k, v in data.items() if k in keep}

class ContactFilter(django_filters.FilterSet):
    """Фильтры API списков контактов"""
    organization = django_filters.CharFilter(field_name='organization__name', lookup_expr='exact')
    organization__contains = django_filters.CharFilter(field_name='organization__name', lookup_expr='contains')
    
    type = django_filters.CharFilter(field_name='type', lookup_expr='exact')
    
    deartment = django_filters.CharFilter(method='filter_department')
    deartment__contains = django_filters.CharFilter(method='filter_department_contains')
    
    manager = django_filters.BooleanFilter(method='filter_manager')

    class Meta:
        model = Contact
        fields = [] 
    
    def filter_department(self, queryset, name, value):
        if not self.data.get("type") or self.data.get("type") != "department":
            return queryset
        return queryset.filter(deartment__name=value)

    def filter_department_contains(self, queryset, name, value):
        if not self.data.get("type") or self.data.get("type") != "department":
            return queryset
        return queryset.filter(deartment__name__contains=value)
    
    def filter_manager(self, queryset, name, value):
        if not self.data.get("type") or self.data.get("type") != "person":
            return queryset
        return queryset.filter(manager=value)

class ContactViewSet(viewsets.ModelViewSet):
    """Viewset контактов, только чтение списка с учетом фильтров"""
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    filterset_class = ContactFilter
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    filter_backends = [DjangoFilterBackend]

    @action(detail=False, methods=["post"], url_path="add")
    def add_contact(self, request):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # вытаскиваем INN из валидированных данных            
        try:
            org = Organization.objects.filter(inn=request.data["organization"]).first()
        except Exception as e:
            return Response(
                {"error": f"Не удалось найти огранизацию контакта"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # создаём контакт с найденной организацией
        contact = Contact.objects.create(
            organization=org,
            type=serializer.validated_data["type"],
            department_name=serializer.validated_data.get("department_name", ""),
            first_name=serializer.validated_data.get("first_name", ""),
            last_name=serializer.validated_data.get("last_name", ""),
            middle_name=serializer.validated_data.get("middle_name", ""),
            first_name_dat=serializer.validated_data.get("first_name_dat", ""),
            last_name_dat=serializer.validated_data.get("last_name_dat", ""),
            middle_name_dat=serializer.validated_data.get("middle_name_dat", ""),
            position=serializer.validated_data.get("position", ""),
            position_dat=serializer.validated_data.get("position_dat", ""),
            manager=serializer.validated_data.get("manager", False),
            comment=serializer.validated_data.get("comment", ""),
            current=serializer.validated_data.get("current", True),
        )

        return Response(ContactSerializer(contact).data, status=status.HTTP_201_CREATED)

class RegionNameSerializer(serializers.ModelSerializer):
    """Сериалайзер для получения списка регионов"""
    class Meta:
        model = Region
        fields = ["name", "code", "is_active"]

class FederalDistrictWithRegionsSerializer(serializers.ModelSerializer):
    """Сериалайзер для получения списка федеральный округов"""
    region = RegionNameSerializer(many=True, read_only=True)
    class Meta:
        model = FederalDistrict
        fields = ["name", "region"]

class OrganizationTypeSerializer(serializers.ModelSerializer):
    """Сериалайзер для получения списка типов организаций"""
    class Meta:
        model = OrganizationType
        fields = "__all__"

class GetAllSerializer(serializers.Serializer):
    """Сериалайзер для получения списков всех обьектов"""
    federal_districts = FederalDistrictWithRegionsSerializer(many=True, read_only=True)
    regions = RegionNameSerializer(many=True, read_only=True)
    organization_types = OrganizationTypeSerializer(many=True, read_only=True)

class GetAllViewSet(viewsets.ViewSet):
    """Вьюсет для определния, списки каких моделей вывести"""
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def list(self, request):
        return Response(
            {
                "message": "Выберите конкретный эндпоинт:",
                "endpoints": {
                    "Регионы": "/contacts/api/get_all/region/",
                    "Типы организаций": "/contacts/api/get_all/organization_type/",
                    "Федеральные округа": "/contacts/api/get_all/fed_district/",
                },
            }
        )
    
    @action(detail=False, methods=["get"], url_path="region")
    def regions(self, request):
        regions = Region.objects.all()
        serializer = RegionNameSerializer(regions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="organization_type")
    def organization_types(self, request):
        org_types = OrganizationType.objects.all()
        serializer = OrganizationTypeSerializer(org_types, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=["get"], url_path="fed_district")
    def federal_districts(self, request):
        districts = FederalDistrict.objects.all()
        serializer = FederalDistrictWithRegionsSerializer(districts, many=True)
        return Response(serializer.data)
